#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
pypi_audit_report.py
Generate a CSV and/or Excel report from one or more pypi_audit JSON outputs.

- Robust JSON ingestion: single object, list of objects, {"results":[...]}, JSON Lines,
  or multiple concatenated top-level JSON values.
- Columns: package, score_total_good, health_percent, risk_level, indicators
  (indicators = semicolon-joined "highlights" from the audit JSON).
- Excel export (optional): adds header styling, autofilter, frozen header,
  and risk-level color coding:
      Critical -> Red
      High     -> Orange
      Medium   -> Yellow
      Low      -> Green
      Very Low -> Blue

Usage:
  python pypi_audit_report.py audit.json -o report.csv
  python pypi_audit_report.py run1.json run2.json --xlsx report.xlsx
  python pypi_audit_report.py merged.json -o report.csv --xlsx report.xlsx
  python pypi_audit_report.py merged.json --stdout

Exit codes:
  0 success
  1 read/parse/write error
"""

import argparse
import csv
import json
import sys
from typing import Iterable, List, Dict, Any, Optional

# Optional Excel dependencies are imported lazily
def _try_import_openpyxl():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        return openpyxl, PatternFill, Font, Alignment
    except Exception as e:
        return None, None, None, None

HEADER = ["package", "score_total_good", "health_percent", "risk_level", "indicators"]

def parse_any_json_stream(text: str) -> List[Any]:
    """
    Extract one or more top-level JSON values from text.
    Strategies:
      1) Full JSON (json.loads)
      2) JSON Lines (one JSON value per non-empty line)
      3) Concatenated JSON values (JSONDecoder.raw_decode loop)
    """
    # 1) Single JSON value
    try:
        val = json.loads(text)
        return [val]
    except json.JSONDecodeError:
        pass

    # 2) JSON Lines
    values = []
    json_lines_candidate = False
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        json_lines_candidate = True
        try:
            values.append(json.loads(s))
        except json.JSONDecodeError:
            values = []
            json_lines_candidate = False
            break
    if json_lines_candidate and values:
        return values

    # 3) Concatenated JSON values
    decoder = json.JSONDecoder()
    idx = 0
    n = len(text)
    values = []
    while idx < n:
        # Skip whitespace
        while idx < n and text[idx].isspace():
            idx += 1
        if idx >= n:
            break
        try:
            val, next_idx = decoder.raw_decode(text, idx)
            values.append(val)
            idx = next_idx
        except json.JSONDecodeError:
            values = []
            break
    if values:
        return values

    raise ValueError("Could not parse JSON content. Ensure valid JSON, JSON Lines, or concatenated JSON.")

def flatten_results(val: Any) -> Iterable[Dict[str, Any]]:
    """
    Yield audit result dicts from shapes:
      - dict result (has 'package' or info.name)
      - list of dicts
      - {'results': [ ... ]}
    """
    if isinstance(val, dict):
        if "package" in val or ("info" in val and isinstance(val["info"], dict) and "name" in val["info"]):
            yield val
        elif "results" in val and isinstance(val["results"], list):
            for item in val["results"]:
                if isinstance(item, dict):
                    yield item
    elif isinstance(val, list):
        for item in val:
            if isinstance(item, dict):
                yield item

def extract_row(rec: Dict[str, Any]) -> Optional[List[Any]]:
    """
    Map a single audit record to CSV/Excel row.
    Skips records that contain an 'error' key.
    """
    if "error" in rec:
        return None
    package = rec.get("package") or (rec.get("info", {}) or {}).get("name") or ""
    score_total_good = rec.get("score_total_good")
    health_percent = rec.get("health_percent")
    risk_level = rec.get("risk_level")
    highlights = rec.get("highlights") or []
    indicators = "; ".join(h for h in highlights if isinstance(h, str))
    return [package, score_total_good, health_percent, risk_level, indicators]

def collect_rows_from_text(text: str) -> List[List[Any]]:
    rows: List[List[Any]] = []
    for val in parse_any_json_stream(text):
        for rec in flatten_results(val):
            row = extract_row(rec)
            if row is not None:
                rows.append(row)
    return rows

def write_csv(path: str, rows: List[List[Any]]) -> None:
    try:
        with open(path, "w", newline="", encoding="utf-8") as f:
            w = csv.writer(f)
            w.writerow(HEADER)
            w.writerows(rows)
    except OSError as e:
        print(f"ERROR: failed to write '{path}': {e}", file=sys.stderr)
        sys.exit(1)

def write_excel(path: str, rows: List[List[Any]]) -> None:
    openpyxl, PatternFill, Font, Alignment = _try_import_openpyxl()
    if openpyxl is None:
        print("WARNING: openpyxl not available; skipping Excel export.", file=sys.stderr)
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PyPI Audit Report"

    # Header
    ws.append(HEADER)
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")
    for col_idx in range(1, len(HEADER) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = header_align

    # Data
    for row in rows:
        ws.append(row)

    # Auto width (simple heuristic)
    for col_idx, col_name in enumerate(HEADER, start=1):
        max_len = len(col_name)
        for row_idx in range(2, ws.max_row + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 2, 80)

    # Freeze header & add filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Risk-level coloring on the "risk_level" column (4th column)
    # Colors chosen to be readable on white background
    fills = {
        "Critical": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # light red
        "High":     PatternFill(start_color="FFEB9C", end_color="FFBB66", fill_type="solid"),  # orange-ish (approx)
        "Medium":   PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # light yellow
        "Low":      PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # light green
        "Very Low": PatternFill(start_color="C6D9F1", end_color="C6D9F1", fill_type="solid"),  # light blue
    }
    risk_col = 4
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=risk_col)
        val = (cell.value or "").strip()
        fill = None
        # Normalize common variants
        if isinstance(val, str):
            low_val = val.lower()
            if low_val == "critical":
                fill = fills["Critical"]
            elif low_val == "high":
                fill = fills["High"]
            elif low_val == "medium":
                fill = fills["Medium"]
            elif low_val == "low":
                fill = fills["Low"]
            elif low_val in ("very low", "very_low", "very-low"):
                fill = fills["Very Low"]
        if fill is not None:
            cell.fill = fill

    try:
        wb.save(path)
    except OSError as e:
        print(f"ERROR: failed to write '{path}': {e}", file=sys.stderr)
        sys.exit(1)

def main():
    ap = argparse.ArgumentParser(description="Generate CSV/Excel from pypi_audit JSON outputs (robust merged/concatenated support).")
    ap.add_argument("inputs", nargs="+", help="One or more JSON files (can be merged, JSONL, or concatenated).")
    ap.add_argument("-o", "--out", default=None, help="Output CSV path. If omitted (and --stdout not used), uses report.csv")
    ap.add_argument("--xlsx", default=None, help="Optional Excel output path (e.g., report.xlsx).")
    ap.add_argument("--stdout", action="store_true", help="Write CSV to stdout instead of a file.")
    args = ap.parse_args()

    all_rows: List[List[Any]] = []
    for path in args.inputs:
        try:
            with open(path, "r", encoding="utf-8") as f:
                text = f.read()
        except OSError as e:
            print(f"ERROR: failed to read '{path}': {e}", file=sys.stderr)
            sys.exit(1)
        try:
            rows = collect_rows_from_text(text)
        except Exception as e:
            print(f"ERROR: failed to parse '{path}': {e}", file=sys.stderr)
            sys.exit(1)
        all_rows.extend(rows)

    # Ensure we have at least an output mechanism
    if not args.stdout and not args.out and not args.xlsx:
        # default CSV path
        args.out = "report.csv"

    # CSV to stdout (if requested)
    if args.stdout:
        w = csv.writer(sys.stdout)
        w.writerow(HEADER)
        w.writerows(all_rows)

    # CSV to file
    if args.out:
        write_csv(args.out, all_rows)

    # Excel to file
    if args.xlsx:
        write_excel(args.xlsx, all_rows)

    # Friendly message when writing files
    if (not args.stdout) and (args.out or args.xlsx):
        parts = []
        if args.out:
            parts.append(f"CSV: {args.out}")
        if args.xlsx:
            parts.append(f"Excel: {args.xlsx}")
        print("Wrote " + ", ".join(parts))

if __name__ == "__main__":
    main()
